"""
Loads rows from the Validation_Set_Template.xlsx "Validation Set" tab and compares
scoring output against partner judgment (PRD SS8). This is the layer that was
deliberately deferred until after the prototype's core scoring engine was built.
"""
import openpyxl

CAPABILITY_NAME_TO_CODE = {
    "Assessment & Due Diligence": "ADD",
    "Market Driven Baseline (MDB)": "MDB",
    "Sales Process Optimization (SPO)": "SPO",
    "Embedded Leadership & Managed Services": "EL",
}

# Maps the spreadsheet's partner-facing "Fit Judgment" dropdown values to the scoring
# engine's band names -- the two vocabularies were designed to line up (PRD SS6.5),
# except "Poor Fit / Not a Fit" maps to "Insufficient Evidence" rather than having its
# own band, since the scoring engine treats "no real fit" as an evidence problem.
FIT_JUDGMENT_TO_EXPECTED_BAND = {
    "Strong Fit": "Strong Fit",
    "Probable Fit": "Probable Fit",
    "Possible Fit": "Possible Fit",
    "Poor Fit / Not a Fit": "Insufficient Evidence",
}


def load_validation_rows(xlsx_path: str) -> list:
    """
    Returns a list of dicts, one per real validation row (skips the header row and the
    highlighted example row on row 2). Keys are the sheet's column headers with any
    trailing " *" required-field marker stripped.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Validation Set" not in wb.sheetnames:
        raise ValueError(f"No 'Validation Set' tab found. Sheets present: {wb.sheetnames}")
    ws = wb["Validation Set"]

    raw_headers = [c.value for c in ws[1]]
    headers = [(h or "").replace(" *", "").strip() for h in raw_headers]

    rows = []
    for row_cells in ws.iter_rows(min_row=3, values_only=True):  # row 1 = header, row 2 = example
        row = dict(zip(headers, row_cells))
        company_name = row.get("Company Name")
        judged_capability = row.get("Judged Capability")
        # Require both a company name AND a judged capability to count as a real data row --
        # guards against picking up stray footer text (e.g. "* Required field") that happens
        # to land in column A below the data range.
        if not company_name or not judged_capability:
            continue
        rows.append(row)
    return rows


def evaluate_row(validation_row: dict, scored_result: dict) -> dict:
    """
    Compares one scored company against its validation row.
    Returns a flat dict suitable for a results table.
    """
    partner_capability_name = validation_row.get("Judged Capability")
    partner_fit = validation_row.get("Fit Judgment")
    partner_code = CAPABILITY_NAME_TO_CODE.get(partner_capability_name)
    expected_band = FIT_JUDGMENT_TO_EXPECTED_BAND.get(partner_fit)

    # scored_result["capabilities"] is already sorted best-to-worst (scoring_engine.py)
    top_capability = scored_result["capabilities"][0]
    top1_match = partner_code is not None and top_capability["capability_code"] == partner_code

    matching_capability_record = next(
        (c for c in scored_result["capabilities"] if c["capability_code"] == partner_code),
        None,
    )
    band_match = (
        matching_capability_record is not None
        and expected_band is not None
        and matching_capability_record["band"] == expected_band
    )

    return {
        "company_name": validation_row.get("Company Name"),
        "partner_capability": partner_capability_name,
        "partner_fit": partner_fit,
        "predicted_top_capability": top_capability["capability_name"],
        "predicted_top_score": top_capability["capability_score_final"],
        "predicted_top_band": top_capability["band"],
        "top1_match": top1_match,
        "predicted_band_for_partner_capability": matching_capability_record["band"] if matching_capability_record else None,
        "expected_band": expected_band,
        "band_match": band_match,
        "discount_applied": matching_capability_record["plausibility_discount_applied"] if matching_capability_record else None,
    }


def summarize(evaluations: list) -> dict:
    n = len(evaluations)
    if n == 0:
        return {"n": 0, "top1_accuracy": None, "band_accuracy": None}
    top1_correct = sum(1 for e in evaluations if e["top1_match"])
    band_correct = sum(1 for e in evaluations if e["band_match"])
    return {
        "n": n,
        "top1_accuracy": round(top1_correct / n, 3),
        "band_accuracy": round(band_correct / n, 3),
    }
